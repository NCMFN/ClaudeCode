import os
import sys
import time
import glob
import pandas as pd
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill
from tabulate import tabulate

def consolidate_excel():
    print("Consolidating Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default sheet

    csv_files = sorted(glob.glob('../tables/*.csv'))

    # Sheet 0: README
    ws_readme = wb.create_sheet(title="README")
    ws_readme.append(["Sheet Name", "Description"])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for cell in ws_readme[1]:
        cell.font = header_font
        cell.fill = header_fill

    for i, file in enumerate(csv_files):
        base_name = os.path.basename(file).replace('.csv', '')
        sheet_name = f"TABLE_{str(i+1).zfill(2)}"
        ws = wb.create_sheet(title=sheet_name)

        df = pd.read_csv(file)

        # Write headers
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        ws.freeze_panes = "A2"

        # Write data
        for r in df.values.tolist():
            ws.append(r)

        # Add to README
        ws_readme.append([sheet_name, base_name])

    wb.save('../outputs/dissertation_tables.xlsx')

    # Also save individual xlsx
    for file in csv_files:
        df = pd.read_csv(file)
        base_name = os.path.basename(file).replace('.csv', '.xlsx')
        df.to_excel(f'../tables/xlsx/{base_name}', index=False)

def main():
    experiments = [
        ('exp01_latency_microbenchmark', 'exp01'),
        ('exp02_throughput_comparison', 'exp02'),
        ('exp03_forex_workload', 'exp03'),
        ('exp04_latency_distribution', 'exp04'),
        ('exp05_latency_over_time', 'exp05'),
        ('exp06_dimension_heatmap', 'exp06'),
        ('exp07_decryption_failure', 'exp07'),
        ('exp08_zk_proof_performance', 'exp08'),
        ('exp09_security_parameter_sweep', 'exp09'),
        ('exp10_batch_commitment', 'exp10'),
        ('exp11_adversarial_budget', 'exp11'),
        ('exp12_selective_opening', 'exp12'),
        ('exp13_multi_trader_contention', 'exp13'),
        ('exp14_grover_bound_visualisation', 'exp14'),
        ('exp15_bkz_attack_complexity', 'exp15')
    ]

    summary = []

    with open('../outputs/run_log.txt', 'w') as log:
        for module_name, exp_name in tqdm(experiments, desc="Running Experiments"):
            try:
                mod = __import__(module_name)
                t0 = time.time()
                # Use reduced trials if applicable to save time
                if hasattr(mod, 'run'):
                    if 'n_trials' in mod.run.__code__.co_varnames:
                        tables, figures = mod.run(n_trials=1000)
                    else:
                        tables, figures = mod.run()
                t1 = time.time()
                runtime = t1 - t0
                status = "PASS"
                log.write(f"[{exp_name}] SUCCESS. Runtime: {runtime:.2f}s\n")
            except Exception as e:
                status = f"FAIL ({str(e)})"
                tables, figures = 0, 0
                runtime = 0.0
                log.write(f"[{exp_name}] FAILED: {str(e)}\n")

            summary.append([exp_name, tables, figures, f"{runtime:.2f}", status])

    consolidate_excel()

    print("\n--- FINAL SUMMARY ---")
    print(tabulate(summary, headers=["Experiment", "Tables", "Figures", "Runtime (s)", "Status"], tablefmt="grid"))

if __name__ == '__main__':
    main()
